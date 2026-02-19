"""
SUPPORT STARTER AI - ANALYTICS EXPORT
=====================================
Export metrics and conversation data to various formats

Features:
- Export to CSV/Excel for analysis
- Export to Google Sheets for live dashboards
- Generate daily/weekly/monthly reports
- Conversation analytics
"""

import os
import csv
import json
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
from pathlib import Path
from dataclasses import dataclass, field
from collections import defaultdict


@dataclass
class ConversationMetrics:
    """Metrics for a single conversation"""
    session_id: str
    started_at: str
    ended_at: Optional[str]
    message_count: int
    customer_provided_name: bool
    customer_provided_email: bool
    customer_provided_phone: bool
    escalated: bool
    status: str
    final_intent: Optional[str] = None
    final_sentiment: Optional[str] = None
    final_lead_score: Optional[int] = None


@dataclass
class DailyMetrics:
    """Aggregated metrics for a day"""
    date: str
    total_conversations: int
    total_messages: int
    escalated_conversations: int
    resolved_conversations: int
    avg_lead_score: float
    sentiment_distribution: Dict[str, int]
    top_intents: List[tuple]  # (intent, count)
    peak_hour: Optional[int] = None


class AnalyticsEngine:
    """
    Generate analytics from conversation logs

    Features:
    - Daily/weekly/monthly reports
    - Sentiment analysis
    - Intent tracking
    - Lead scoring trends
    """

    def __init__(self, log_dir: str = "chat_logs/sessions"):
        self.log_dir = Path(log_dir)

    def scan_conversations(self, days_back: int = 30) -> List[Dict]:
        """
        Scan conversation logs

        Args:
            days_back: How many days back to scan

        Returns:
            List of conversation data
        """
        conversations = []
        cutoff_date = datetime.now() - timedelta(days=days_back)

        for file_path in self.log_dir.glob("*.json"):
            try:
                # Check file modification time
                file_mtime = datetime.fromtimestamp(os.path.getmtime(file_path))
                if file_mtime < cutoff_date:
                    continue

                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                conversations.append(data)

            except Exception as e:
                print(f"Error reading {file_path}: {e}")

        return conversations

    def calculate_daily_metrics(self, date: str) -> DailyMetrics:
        """
        Calculate metrics for a specific date

        Args:
            date: Date in YYYY-MM-DD format

        Returns:
            DailyMetrics object
        """
        conversations = self.scan_conversations(days_back=365)

        # Filter conversations for the specific date
        day_convs = [c for c in conversations if c.get("started_at", "").startswith(date)]

        if not day_convs:
            return DailyMetrics(
                date=date,
                total_conversations=0,
                total_messages=0,
                escalated_conversations=0,
                resolved_conversations=0,
                avg_lead_score=0.0,
                sentiment_distribution={},
                top_intents=[]
            )

        total_messages = sum(len(c.get("messages", [])) for c in day_convs)
        escalated = sum(1 for c in day_convs if c.get("escalated", False))
        resolved = sum(1 for c in day_convs if c.get("status") == "resolved")

        # Calculate sentiment distribution
        sentiment_counts = defaultdict(int)
        lead_scores = []

        for conv in day_convs:
            # Get last message metadata for sentiment/lead score
            messages = conv.get("messages", [])
            if messages:
                last_msg = messages[-1]
                sentiment = last_msg.get("metadata", {}).get("sentiment")
                lead_score = last_msg.get("metadata", {}).get("lead_score")

                if sentiment:
                    sentiment_counts[sentiment] += 1
                if lead_score is not None:
                    lead_scores.append(lead_score)

        avg_lead_score = sum(lead_scores) / len(lead_scores) if lead_scores else 0

        # Get top intents
        intent_counts = defaultdict(int)
        for conv in day_convs:
            for msg in conv.get("messages", []):
                intent = msg.get("metadata", {}).get("intent")
                if intent and intent not in ["unknown", "greeting", "goodbye"]:
                    intent_counts[intent] += 1

        top_intents = sorted(intent_counts.items(), key=lambda x: x[1], reverse=True)[:5]

        return DailyMetrics(
            date=date,
            total_conversations=len(day_convs),
            total_messages=total_messages,
            escalated_conversations=escalated,
            resolved_conversations=resolved,
            avg_lead_score=avg_lead_score,
            sentiment_distribution=dict(sentiment_counts),
            top_intents=top_intents
        )

    def generate_weekly_report(self, week_start: str = None) -> Dict[str, Any]:
        """
        Generate weekly analytics report

        Args:
            week_start: Start date in YYYY-MM-DD (defaults to current week)

        Returns:
            Weekly report dict
        """
        if week_start is None:
            # Get Monday of current week
            today = datetime.now()
            week_start = (today - timedelta(days=today.weekday())).strftime("%Y-%m-%d")

        daily_metrics = []
        for i in range(7):
            date = datetime.strptime(week_start, "%Y-%m-%d") + timedelta(days=i)
            date_str = date.strftime("%Y-%m-%d")
            daily_metrics.append(self.calculate_daily_metrics(date_str))

        # Aggregate week metrics
        total_conversations = sum(d.total_conversations for d in daily_metrics)
        total_messages = sum(d.total_messages for d in daily_metrics)
        total_escalated = sum(d.escalated_conversations for d in daily_metrics)
        total_resolved = sum(d.resolved_conversations for d in daily_metrics)

        # Average lead score across week
        all_lead_scores = []
        for d in daily_metrics:
            if d.avg_lead_score > 0:
                all_lead_scores.append(d.avg_lead_score)
        avg_lead_score = sum(all_lead_scores) / len(all_lead_scores) if all_lead_scores else 0

        # Top intents for the week
        intent_counts = defaultdict(int)
        for d in daily_metrics:
            for intent, count in d.top_intents:
                intent_counts[intent] += count

        top_intents = sorted(intent_counts.items(), key=lambda x: x[1], reverse=True)[:10]

        return {
            "week_start": week_start,
            "week_end": (datetime.strptime(week_start, "%Y-%m-%d") + timedelta(days=6)).strftime("%Y-%m-%d"),
            "summary": {
                "total_conversations": total_conversations,
                "total_messages": total_messages,
                "escalated_conversations": total_escalated,
                "resolved_conversations": total_resolved,
                "escalation_rate": total_escalated / total_conversations if total_conversations > 0 else 0,
                "resolution_rate": total_resolved / total_conversations if total_conversations > 0 else 0,
                "avg_lead_score": avg_lead_score
            },
            "top_intents": top_intents,
            "daily_breakdown": [
                {
                    "date": d.date,
                    "conversations": d.total_conversations,
                    "escalated": d.escalated_conversations
                }
                for d in daily_metrics
            ]
        }


class AnalyticsExporter:
    """
    Export analytics to various formats

    Supports:
    - CSV files
    - Excel files
    - JSON reports
    - Google Sheets (if configured)
    """

    def __init__(self, analytics: AnalyticsEngine = None):
        self.analytics = analytics or AnalyticsEngine()

    def export_to_csv(self, output_path: str = "analytics_export.csv") -> str:
        """
        Export conversation data to CSV

        Args:
            output_path: Path to output CSV file

        Returns:
            Path to created file
        """
        conversations = self.analytics.scan_conversations(days_back=30)

        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Header
            writer.writerow([
                "Session ID", "Started At", "Ended At", "Message Count",
                "Escalated", "Status", "Customer Name", "Customer Email",
                "Final Intent", "Final Sentiment", "Final Lead Score"
            ])

            # Rows
            for conv in conversations:
                # Get final metrics from last message
                messages = conv.get("messages", [])
                final_intent = None
                final_sentiment = None
                final_lead_score = None

                if messages:
                    last_msg = messages[-1]
                    metadata = last_msg.get("metadata", {})
                    final_intent = metadata.get("intent")
                    final_sentiment = metadata.get("sentiment")
                    final_lead_score = metadata.get("lead_score")

                writer.writerow([
                    conv.get("session_id"),
                    conv.get("started_at"),
                    conv.get("ended_at", ""),
                    len(messages),
                    conv.get("escalated", False),
                    conv.get("status", ""),
                    conv.get("customer_name", ""),
                    conv.get("customer_email", ""),
                    final_intent,
                    final_sentiment,
                    final_lead_score
                ])

        return output_path

    def export_to_excel(self, output_path: str = "analytics_export.xlsx") -> str:
        """
        Export analytics to Excel with multiple sheets

        Args:
            output_path: Path to output Excel file

        Returns:
            Path to created file
        """
        try:
            import openpyxl
            from openpyxl import Workbook
        except ImportError:
            print("openpyxl not installed. Run: pip install openpyxl")
            return None

        conversations = self.analytics.scan_conversations(days_back=30)

        wb = Workbook()

        # Sheet 1: Conversations
        ws1 = wb.active
        ws1.title = "Conversations"
        ws1.append([
            "Session ID", "Started At", "Message Count", "Escalated", "Status",
            "Customer Email", "Final Intent", "Final Lead Score"
        ])

        for conv in conversations:
            messages = conv.get("messages", [])
            final_intent = None
            final_lead_score = None

            if messages:
                last_msg = messages[-1]
                metadata = last_msg.get("metadata", {})
                final_intent = metadata.get("intent")
                final_lead_score = metadata.get("lead_score")

            ws1.append([
                conv.get("session_id"),
                conv.get("started_at"),
                len(messages),
                conv.get("escalated", False),
                conv.get("status", ""),
                conv.get("customer_email", ""),
                final_intent,
                final_lead_score
            ])

        # Sheet 2: Daily Metrics
        ws2 = wb.create_sheet("Daily Metrics")
        ws2.append(["Date", "Conversations", "Escalated", "Resolved", "Avg Lead Score"])

        # Get last 7 days
        for i in range(7):
            date = datetime.now() - timedelta(days=6-i)
            date_str = date.strftime("%Y-%m-%d")
            metrics = self.analytics.calculate_daily_metrics(date_str)

            ws2.append([
                date_str,
                metrics.total_conversations,
                metrics.escalated_conversations,
                metrics.resolved_conversations,
                round(metrics.avg_lead_score, 2)
            ])

        # Sheet 3: Top Intents
        ws3 = wb.create_sheet("Top Intents")
        ws3.append(["Intent", "Count"])

        # Aggregate intents across all conversations
        intent_counts = defaultdict(int)
        for conv in conversations:
            for msg in conv.get("messages", []):
                intent = msg.get("metadata", {}).get("intent")
                if intent and intent not in ["unknown", "greeting", "goodbye"]:
                    intent_counts[intent] += 1

        for intent, count in sorted(intent_counts.items(), key=lambda x: x[1], reverse=True)[:20]:
            ws3.append([intent, count])

        wb.save(output_path)
        return output_path

    def export_to_sheets(self, sheet_id: str, sheet_name: str = "Analytics") -> Dict[str, Any]:
        """
        Export analytics to Google Sheets

        Args:
            sheet_id: Google Sheet ID
            sheet_name: Name of the sheet tab

        Returns:
            Export result
        """
        try:
            from sheets_admin import SheetsAdmin
        except ImportError:
            return {"error": "Google API not installed"}

        # Generate weekly report
        report = self.analytics.generate_weekly_report()

        # For now, return the report data
        # Full implementation would append to the sheet
        return {
            "status": "report_generated",
            "sheet_id": sheet_id,
            "sheet_name": sheet_name,
            "data": report
        }

    def generate_dashboard_html(self, output_path: str = "dashboard.html") -> str:
        """
        Generate a simple HTML dashboard

        Args:
            output_path: Path to output HTML file

        Returns:
            Path to created file
        """
        conversations = self.analytics.scan_conversations(days_back=7)

        # Calculate metrics
        total_conversations = len(conversations)
        escalated = sum(1 for c in conversations if c.get("escalated", False))
        resolved = sum(1 for c in conversations if c.get("status") == "resolved")

        # Get top intents
        intent_counts = defaultdict(int)
        for conv in conversations:
            for msg in conv.get("messages", []):
                intent = msg.get("metadata", {}).get("intent")
                if intent and intent not in ["unknown", "greeting", "goodbye"]:
                    intent_counts[intent] += 1

        top_intents = sorted(intent_counts.items(), key=lambda x: x[1], reverse=True)[:5]

        # Generate HTML
        html = f"""<!DOCTYPE html>
<html>
<head>
    <title>Support Starter Analytics</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 40px; background: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; }}
        h1 {{ color: #333; }}
        .metrics {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 20px; margin: 30px 0; }}
        .metric-card {{ background: #f8f9fa; padding: 20px; border-radius: 8px; text-align: center; }}
        .metric-value {{ font-size: 36px; font-weight: bold; color: #007bff; }}
        .metric-label {{ color: #666; font-size: 14px; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }}
        th {{ background: #f8f9fa; }}
        .timestamp {{ color: #999; font-size: 12px; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🤖 Support Starter Analytics</h1>
        <p class="timestamp">Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>

        <div class="metrics">
            <div class="metric-card">
                <div class="metric-value">{total_conversations}</div>
                <div class="metric-label">Conversations (7 days)</div>
            </div>
            <div class="metric-card">
                <div class="metric-value">{resolved}</div>
                <div class="metric-label">Resolved</div>
            </div>
            <div class="metric-card">
                <div class="metric-value">{escalated}</div>
                <div class="metric-label">Escalated</div>
            </div>
        </div>

        <h2>Top Intents</h2>
        <table>
            <tr><th>Intent</th><th>Count</th></tr>
            {"".join(f"<tr><td>{intent}</td><td>{count}</td></tr>" for intent, count in top_intents)}
        </table>
    </div>
</body>
</html>"""

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html)

        return output_path


# Convenience functions
def export_analytics(format: str = "csv", output_path: str = None) -> str:
    """
    Quick export function

    Args:
        format: Output format (csv, excel, html)
        output_path: Optional output path

    Returns:
        Path to exported file
    """
    exporter = AnalyticsExporter()

    if format == "csv":
        output_path = output_path or "analytics_export.csv"
        return exporter.export_to_csv(output_path)
    elif format == "excel":
        output_path = output_path or "analytics_export.xlsx"
        return exporter.export_to_excel(output_path)
    elif format == "html":
        output_path = output_path or "dashboard.html"
        return exporter.generate_dashboard_html(output_path)
    else:
        raise ValueError(f"Unknown format: {format}")


if __name__ == "__main__":
    print("=" * 60)
    print("ANALYTICS EXPORT - TEST")
    print("=" * 60)

    # Test analytics
    analytics = AnalyticsEngine()

    # Generate weekly report
    report = analytics.generate_weekly_report()
    print("\n--- Weekly Report ---")
    print(json.dumps(report, indent=2, ensure_ascii=False))

    # Test exports
    print("\n--- Testing Exports ---")
    exporter = AnalyticsExporter()

    # Export to CSV
    csv_path = exporter.export_to_csv("test_analytics.csv")
    print(f"CSV exported to: {csv_path}")

    # Generate dashboard
    html_path = exporter.generate_dashboard_html("test_dashboard.html")
    print(f"Dashboard generated: {html_path}")
